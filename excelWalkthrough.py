from openpyxl import load_workbook
import re


def get_stack(cell_vall):
    return cell_vall[:3]


def get_num_stacks(codes_list):
    nums = {}
    for stages in range(1, 4*5+1):
        nums[stages] = 0
    for codes in codes_list.values():
        nums[len(codes)] += 1
    return nums


regalFilter = re.compile('^[A-Z][0-9]{2}-[0-9]{2}-[0-9]{2}$')

wb = load_workbook('regalyVB.xlsx')
ws = wb.active

print('test', ws['DD32'].value)
print(f'prvni {ws.cell(row=32,column=23).value}')
print(f'posledni {ws.cell(row=195,column=25*4+23).value}')

empty = 0
codes = 0

codes_list = {}

for column in ws.iter_cols(min_row=32,
                           max_row=216,
                           min_col=23,
                           max_col=25*4+23):
    for cell in column:
        code_line = cell.value
        if isinstance(code_line, str) and regalFilter.match(code_line):
            codes += 1
            code_line = cell.value
            code_key = get_stack(code_line)
            if code_key in codes_list:
                codes_list[code_key].append(code_line)
            else:
                codes_list[code_key] = [code_line]
        else:
            empty += 1

#print("pocty: \n", list(map(lambda x : x.value >0,get_num_stacks(codes_list))))

print(f'prazdne: {empty} coduu: {codes}')

import pickle

pickle.dump(codes_list, open('dumped_codes','wb'))