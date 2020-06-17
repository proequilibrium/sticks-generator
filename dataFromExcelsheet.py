import pickle
from openpyxl import load_workbook
import re


def get_stack(cell_vall):
    return cell_vall[:3]


def get_num_stacks(codes_list):
    nums = {}
    for stages in range(1, 4*5+1):
        nums[stages] = 0
    for codes in codes_list.values():
        nums[len(codes)] += 1
    return nums


def get_lost_values_list(excelSheet='dodelat.xlsx', lines=1):
    regalFilter = re.compile('^[A-Z][0-9]{2}-[0-9]{2}-[0-9]{2}$')

    wb = load_workbook(excelSheet)
    ws = wb.active

    empty = 0
    codes = 0

    codes_list = []

    for column in ws.iter_cols(min_row=1,
                               max_row=lines,
                               min_col=1,
                               max_col=1):

        for cell in column:
            code_line = cell.value
            if isinstance(code_line, str) and regalFilter.match(code_line):
                codes += 1
                code_line = cell.value
                code_key = get_stack(code_line)
                codes_list.append(code_line)
            else:
                empty += 1
    print(f'prazdne: {empty} coduu: {codes}')
    return codes_list
#print("pocty: \n", list(map(lambda x : x.value >0,get_num_stacks(codes_list))))


if __name__ == "__main__":
    print(get_lost_values_list(excelSheet='dodelat_test.xlsx', lines=5))
