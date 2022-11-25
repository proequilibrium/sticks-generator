import argparse

from openpyxl import load_workbook
import re
import pprint

MAPPING = {
    "P": ["A","B","E","F"],
    "R": ["C","D","G","H"]
    }

def get_stack(cell_vall):
    return cell_vall[:3]


def get_num_stacks(codes_list):
    nums = {}
    for stages in range(1, 4 * 5 + 1):
        nums[stages] = 0
    for codes in codes_list.values():
        nums[len(codes)] += 1
    return nums

def get_batched_codes(codes_list):
    batched_codes = {}
    for code in codes_list:
        stage_key = get_stack(code)
        if stage_key in batched_codes:
            batched_codes[stage_key].append(code)
        else:
            batched_codes[stage_key] = [code]
    return batched_codes

def prepair_data_for_generation(excelSheet="stow/regalyVB.xlsx") -> dict:
    codes_list = get_values_from_xlsx(excelSheet)
    codes_list = map_codes(codes_list)
    batched_codes = get_batched_codes(codes_list)
    return batched_codes

def get_values_from_xlsx(excelSheet="stow/regalyVB.xlsx"):
    regalFilter = re.compile("^[A-Z][0-9]{2}-[0-9]{2}-[0-9]{2}$")

    wb = load_workbook(excelSheet)
    ws = wb.active

    empty = 0
    codes = 0

    codes_list = []
    print(ws['X40'].value, f"{ws.max_column=}, {ws.max_row=}")
    for column in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):

        for cell in column:
            code_line = cell.value
            if isinstance(code_line, str) and regalFilter.match(code_line):
                codes += 1
                code_line = cell.value
                code_key = get_stack(code_line)
                codes_list.append(code_line)
            else:
                empty += 1
    pprint.pprint(f"{empty=} {codes=}")
    return codes_list
    # print("pocty: \n", list(map(lambda x : x.value >0,get_num_stacks(codes_list))))

def map_codes(codes_list):
    maped = []
    for from_value, to_values in MAPPING.items():
        for to_value in to_values:
            new_maped = [ code.replace(from_value, to_value) for code in codes_list if code[0] in from_value]
            maped.extend(new_maped)
    return maped

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("-f", "--file", type=str, default="stow/regalyVB.xlsx", help="Excel file")
    args = parser.parse_args()

    codes = get_values_from_xlsx(excelSheet=args.file)
    maped_codes=map_codes(codes)
    print(maped_codes)
    print(len(maped_codes))
