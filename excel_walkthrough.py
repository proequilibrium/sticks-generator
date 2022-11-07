from openpyxl import load_workbook
import re
import pickle


def get_stack(cell_vall):
    return cell_vall[:3]


def get_num_stacks(codes_list):
    nums = {}
    for stages in range(1, 4 * 5 + 1):
        nums[stages] = 0
    for codes in codes_list.values():
        nums[len(codes)] += 1
    return nums


rep = {"P": "T", "Q": "U", "R": "V", "S": "X"}  # define desired replacements here

# # use these three lines to do the replacement
rep = dict((re.escape(k), v) for k, v in rep.items())
# Python 3 renamed dict.iteritems to dict.items so use rep.items() for latest versions
code_pattern = re.compile("|".join(rep.keys()))


def replace_all(code: str):
    text = code_pattern.sub(lambda m: rep[re.escape(m.group(0))], code)
    return text


# text = pattern.sub(lambda m: rep[re.escape(m.group(0))], text)

regalFilter = re.compile("^[A-Z][0-9]{2}-[0-9]{2}-[0-9]{2}$")


wb = load_workbook("regalyVB.xlsx")
ws = wb.active

print("test", ws["DS55"].value)
print(f"prvni {ws.cell(row=32,column=23).value}")
print(f"posledni {ws.cell(row=195,column=25*4+23).value}")

empty = 0

codes_list = []
codes_by_code = {}
# for column in ws.iter_cols(min_row=32,
#                            max_row=216,
#                            min_col=23,
#    max_col=25*4+23):
for column in ws.iter_cols(
    min_row=32,
    max_row=64,
    min_col=(ord("x") - ord("a")),
    max_col=(ord("z") - ord("a") + 1) * 4 + (ord("s") - ord("a")) + 1,
):
    for cell in column:
        code_line = cell.value
        if isinstance(code_line, str) and regalFilter.match(code_line):
            code_line = cell.value
            code_line = replace_all(code_line)
            code_key = get_stack(code_line)
            codes_list.append(code_line)
            # print(code_line)
            if code_key in codes_list:
                codes_by_code[code_key].append(code_line)
            else:
                codes_by_code[code_key] = [code_line]
        else:
            empty += 1

# print("pocty: \n", list(map(lambda x :
# x.value >0,get_num_stacks(codes_list))))

print(f"prazdne: {empty} codu: {len(codes_list)}")
new_codes_list = sorted(codes_list, reverse=True)

print(f"{new_codes_list}")

with open("./dumped_codes", "w") as dump_file:
    for code in new_codes_list:
        dump_file.write(f"{code}\n")
